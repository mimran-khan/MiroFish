"""
File parsing utilities.
Extracts text from PDF, Markdown, TXT, CSV, and XLSX.
"""

import csv
import io
import os
from pathlib import Path
from typing import List, Optional


def _read_text_with_fallback(file_path: str) -> str:
    """
    Read a text file; fall back through encodings if UTF-8 fails.

    Strategy:
    1. Try UTF-8
    2. Detect with charset_normalizer
    3. Fall back to chardet
    4. Final fallback: UTF-8 with errors='replace'

    Args:
        file_path: Path to the file

    Returns:
        Decoded text
    """
    data = Path(file_path).read_bytes()
    
    try:
        return data.decode('utf-8')
    except UnicodeDecodeError:
        pass
    
    encoding = None
    try:
        from charset_normalizer import from_bytes
        best = from_bytes(data).best()
        if best and best.encoding:
            encoding = best.encoding
    except Exception:
        pass
    
    if not encoding:
        try:
            import chardet
            result = chardet.detect(data)
            encoding = result.get('encoding') if result else None
        except Exception:
            pass
    
    if not encoding:
        encoding = 'utf-8'
    
    return data.decode(encoding, errors='replace')


class FileParser:
    """Extract plain text from supported file types."""
    
    SUPPORTED_EXTENSIONS = {'.pdf', '.md', '.markdown', '.txt', '.csv', '.xlsx'}
    
    @classmethod
    def extract_text(cls, file_path: str) -> str:
        """
        Extract text from a single file.

        Args:
            file_path: Path to the file

        Returns:
            Extracted text
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        suffix = path.suffix.lower()
        
        if suffix not in cls.SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported file type: {suffix}")
        
        if suffix == '.pdf':
            return cls._extract_from_pdf(file_path)
        elif suffix in {'.md', '.markdown'}:
            return cls._extract_from_md(file_path)
        elif suffix == '.txt':
            return cls._extract_from_txt(file_path)
        elif suffix == '.csv':
            return cls._extract_from_csv(file_path)
        elif suffix == '.xlsx':
            return cls._extract_from_xlsx(file_path)
        
        raise ValueError(f"Cannot handle file type: {suffix}")
    
    @staticmethod
    def _extract_from_pdf(file_path: str) -> str:
        """Extract text from PDF."""
        try:
            import fitz  # PyMuPDF
        except ImportError:
            raise ImportError("PyMuPDF is required: pip install PyMuPDF")
        
        text_parts = []
        with fitz.open(file_path) as doc:
            for page in doc:
                text = page.get_text()
                if text.strip():
                    text_parts.append(text)
        
        return "\n\n".join(text_parts)
    
    @staticmethod
    def _extract_from_md(file_path: str) -> str:
        """Read Markdown with encoding detection."""
        return _read_text_with_fallback(file_path)
    
    @staticmethod
    def _extract_from_txt(file_path: str) -> str:
        """Read plain text with encoding detection."""
        return _read_text_with_fallback(file_path)
    
    @staticmethod
    def _extract_from_csv(file_path: str) -> str:
        """Turn CSV rows into key: value style lines."""
        raw = _read_text_with_fallback(file_path)
        reader = csv.DictReader(io.StringIO(raw))
        
        if not reader.fieldnames:
            return raw
        
        rows = []
        for row in reader:
            parts = [f"{col}: {val.strip()}" for col, val in row.items() if val and val.strip()]
            if parts:
                rows.append(" | ".join(parts))
        
        if not rows:
            return raw
        
        header = "Columns: " + ", ".join(reader.fieldnames)
        return header + "\n\n" + "\n".join(rows)
    
    @staticmethod
    def _extract_from_xlsx(file_path: str) -> str:
        """Turn XLSX rows into key: value style lines per sheet."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required: pip install openpyxl")
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        all_sheets = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                continue
            
            headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(sheet_rows[0])]
            lines = []
            for row in sheet_rows[1:]:
                parts = []
                for col_name, val in zip(headers, row):
                    if val is not None:
                        parts.append(f"{col_name}: {str(val).strip()}")
                if parts:
                    lines.append(" | ".join(parts))
            
            if lines:
                sheet_header = f"=== Sheet: {sheet_name} ===\nColumns: {', '.join(headers)}"
                all_sheets.append(sheet_header + "\n\n" + "\n".join(lines))
        
        wb.close()
        return "\n\n".join(all_sheets) if all_sheets else ""
    
    @classmethod
    def extract_from_multiple(cls, file_paths: List[str]) -> str:
        """
        Extract and concatenate multiple files.

        Args:
            file_paths: List of paths

        Returns:
            Merged text with per-file headers
        """
        all_texts = []
        
        for i, file_path in enumerate(file_paths, 1):
            try:
                text = cls.extract_text(file_path)
                filename = Path(file_path).name
                all_texts.append(f"=== Document {i}: {filename} ===\n{text}")
            except Exception as e:
                all_texts.append(
                    f"=== Document {i}: {file_path} (extraction failed: {str(e)}) ==="
                )
        
        return "\n\n".join(all_texts)


def split_text_into_chunks(
    text: str, 
    chunk_size: int = 500, 
    overlap: int = 50
) -> List[str]:
    """
    Split text into overlapping chunks.

    Args:
        text: Source text
        chunk_size: Target chunk length in characters
        overlap: Overlap between consecutive chunks

    Returns:
        List of chunk strings
    """
    if len(text) <= chunk_size:
        return [text] if text.strip() else []
    
    chunks = []
    start = 0
    
    while start < len(text):
        end = start + chunk_size
        
        if end < len(text):
            # Prefer breaking at sentence boundaries (CJK + Latin)
            for sep in ['\u3002', '\uff01', '\uff1f', '.\n', '!\n', '?\n', '\n\n', '. ', '! ', '? ']:
                last_sep = text[start:end].rfind(sep)
                if last_sep != -1 and last_sep > chunk_size * 0.3:
                    end = start + last_sep + len(sep)
                    break
        
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        
        start = end - overlap if end < len(text) else len(text)
    
    return chunks
